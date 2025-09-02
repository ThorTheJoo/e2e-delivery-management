#!/usr/bin/env python3
"""
Excel Analyzer - Comprehensive CLI tool for analyzing Excel (.xlsx) and macro-enabled Excel (.xlsm) files

This tool provides deep analysis of Excel files including:
- Structural analysis (sheets, cells, ranges)
- Content analysis (data, formulas, validation)
- Formatting analysis (styles, colors, fonts)
- VBA macro analysis (code extraction, security analysis)
- Markdown report generation for documentation and reuse

Usage:
    python excel_analyzer.py analyze file.xlsx
    python excel_analyzer.py analyze file.xlsm --include-vba --output-format markdown
    python excel_analyzer.py compare file1.xlsx file2.xlsm
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import warnings

# Core libraries
import openpyxl
import pandas as pd
from oletools.olevba import VBA_Parser

# Suppress openpyxl warnings for cleaner output
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

class ExcelAnalyzer:
    """Main Excel analysis engine"""
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.file_name = self.file_path.name
        self.is_macro_enabled = self.file_path.suffix.lower() == '.xlsm'
        self.analysis_results = {}
        
    def analyze(self, include_vba: bool = True, include_formatting: bool = True) -> Dict[str, Any]:
        """Perform comprehensive analysis of the Excel file"""
        print(f"🔍 Analyzing {self.file_name}...")
        
        results = {
            "file_info": self._get_file_info(),
            "metadata": self._analyze_metadata(),
            "structure": self._analyze_structure(),
            "content": self._analyze_content(),
        }
        
        if include_formatting:
            print("  📊 Analyzing formatting...")
            results["formatting"] = self._analyze_formatting()
            
        if include_vba and self.is_macro_enabled:
            print("  🔧 Analyzing VBA macros...")
            results["vba_analysis"] = self._analyze_vba()
        elif include_vba:
            results["vba_analysis"] = {"message": "File is not macro-enabled"}
            
        self.analysis_results = results
        return results
    
    def _get_file_info(self) -> Dict[str, Any]:
        """Get basic file information"""
        stat = self.file_path.stat()
        return {
            "file_name": self.file_name,
            "file_path": str(self.file_path),
            "file_size": stat.st_size,
            "file_size_mb": round(stat.st_size / (1024 * 1024), 2),
            "is_macro_enabled": self.is_macro_enabled,
            "last_modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
            "analysis_timestamp": datetime.now().isoformat()
        }
    
    def _analyze_metadata(self) -> Dict[str, Any]:
        """Analyze workbook metadata using openpyxl"""
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            
            metadata = {
                "creator": wb.properties.creator,
                "title": wb.properties.title,
                "subject": wb.properties.subject,
                "description": wb.properties.description,
                "keywords": wb.properties.keywords,
                "category": wb.properties.category,
                "created": str(wb.properties.created) if wb.properties.created else None,
                "modified": str(wb.properties.modified) if wb.properties.modified else None,
                "last_modified_by": wb.properties.lastModifiedBy,
                "revision": wb.properties.revision,
                "version": wb.properties.version,
                "defined_names": list(wb.defined_names.keys()) if wb.defined_names else []
            }
            
            wb.close()
            return metadata
            
        except Exception as e:
            return {"error": f"Failed to analyze metadata: {str(e)}"}
    
    def _analyze_structure(self) -> Dict[str, Any]:
        """Analyze workbook structure using pandas and openpyxl"""
        try:
            # Use pandas for efficient sheet enumeration
            all_sheets = pd.read_excel(self.file_path, sheet_name=None, engine='openpyxl', nrows=0)
            
            # Use openpyxl for detailed sheet analysis
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            
            structure = {
                "sheet_count": len(all_sheets),
                "sheet_names": list(all_sheets.keys()),
                "sheets": {}
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                structure["sheets"][sheet_name] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "dimensions": ws.dimensions,
                    "merged_cells_count": len(ws.merged_cells.ranges) if ws.merged_cells else 0,
                    "has_charts": len(ws._charts) > 0 if hasattr(ws, '_charts') else False,
                    "has_images": len(ws._images) > 0 if hasattr(ws, '_images') else False,
                    "sheet_state": ws.sheet_state,
                    "protection": {
                        "sheet_protected": ws.protection.sheet,
                        "password_protected": bool(ws.protection.password)
                    } if ws.protection else None
                }
            
            wb.close()
            return structure
            
        except Exception as e:
            return {"error": f"Failed to analyze structure: {str(e)}"}
    
    def _analyze_content(self) -> Dict[str, Any]:
        """Analyze content using pandas for efficient data processing"""
        try:
            # Read all sheets with pandas
            all_sheets = pd.read_excel(self.file_path, sheet_name=None, engine='openpyxl')
            
            content = {
                "total_rows": 0,
                "total_columns": 0,
                "sheets": {}
            }
            
            for sheet_name, df in all_sheets.items():
                if not df.empty:
                    # Data analysis
                    sheet_analysis = {
                        "rows": len(df),
                        "columns": len(df.columns),
                        "column_names": list(df.columns),
                        "data_types": {str(k): str(v) for k, v in df.dtypes.to_dict().items()},
                        "null_counts": df.isnull().sum().to_dict(),
                        "non_null_counts": df.count().to_dict(),
                        "has_formulas": False,  # Will be updated by openpyxl analysis
                        "sample_data": df.head(3).fillna("").to_dict() if not df.empty else {}
                    }
                    
                    content["total_rows"] += len(df)
                    content["total_columns"] += len(df.columns)
                else:
                    sheet_analysis = {
                        "rows": 0,
                        "columns": 0,
                        "column_names": [],
                        "data_types": {},
                        "null_counts": {},
                        "non_null_counts": {},
                        "has_formulas": False,
                        "sample_data": {}
                    }
                
                content["sheets"][sheet_name] = sheet_analysis
            
            # Check for formulas using openpyxl
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                has_formulas = False
                formula_count = 0
                
                for row in ws.iter_rows(max_row=min(100, ws.max_row)):  # Sample first 100 rows
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            has_formulas = True
                            formula_count += 1
                
                if sheet_name in content["sheets"]:
                    content["sheets"][sheet_name]["has_formulas"] = has_formulas
                    content["sheets"][sheet_name]["formula_count_sample"] = formula_count
            
            wb.close()
            return content
            
        except Exception as e:
            return {"error": f"Failed to analyze content: {str(e)}"}
    
    def _analyze_formatting(self) -> Dict[str, Any]:
        """Analyze formatting using openpyxl"""
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=False)
            
            formatting = {
                "sheets": {},
                "summary": {
                    "total_styled_cells": 0,
                    "unique_fonts": set(),
                    "unique_colors": set(),
                    "has_conditional_formatting": False
                }
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_formatting = {
                    "styled_cells": 0,
                    "fonts": {},
                    "colors": {},
                    "borders": 0,
                    "conditional_formatting_rules": len(ws.conditional_formatting) if hasattr(ws, 'conditional_formatting') else 0
                }
                
                # Sample formatting from first 50x50 cells
                max_sample_row = min(50, ws.max_row)
                max_sample_col = min(50, ws.max_column)
                
                for row in range(1, max_sample_row + 1):
                    for col in range(1, max_sample_col + 1):
                        cell = ws.cell(row=row, column=col)
                        
                        if cell.font and cell.font.name:
                            font_key = f"{cell.font.name}_{cell.font.size}"
                            sheet_formatting["fonts"][font_key] = sheet_formatting["fonts"].get(font_key, 0) + 1
                            formatting["summary"]["unique_fonts"].add(font_key)
                        
                        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                            color = str(cell.fill.fgColor.rgb) if cell.fill.fgColor.rgb else "auto"
                            sheet_formatting["colors"][color] = sheet_formatting["colors"].get(color, 0) + 1
                            formatting["summary"]["unique_colors"].add(color)
                        
                        if cell.border and any([cell.border.left.style, cell.border.right.style, 
                                              cell.border.top.style, cell.border.bottom.style]):
                            sheet_formatting["borders"] += 1
                        
                        if any([cell.font and cell.font.name, cell.fill and cell.fill.fgColor, 
                               cell.border and cell.border.left.style]):
                            sheet_formatting["styled_cells"] += 1
                            formatting["summary"]["total_styled_cells"] += 1
                
                if sheet_formatting["conditional_formatting_rules"] > 0:
                    formatting["summary"]["has_conditional_formatting"] = True
                
                formatting["sheets"][sheet_name] = sheet_formatting
            
            # Convert sets to lists for JSON serialization
            formatting["summary"]["unique_fonts"] = list(formatting["summary"]["unique_fonts"])
            formatting["summary"]["unique_colors"] = list(formatting["summary"]["unique_colors"])
            
            wb.close()
            return formatting
            
        except Exception as e:
            return {"error": f"Failed to analyze formatting: {str(e)}"}
    
    def _analyze_vba(self) -> Dict[str, Any]:
        """Analyze VBA macros using oletools"""
        if not self.is_macro_enabled:
            return {"message": "File is not macro-enabled"}
        
        try:
            vba_parser = VBA_Parser(str(self.file_path))
            
            vba_analysis = {
                "has_macros": vba_parser.detect_vba_macros(),
                "modules": [],
                "security_analysis": {
                    "suspicious_keywords": [],
                    "auto_exec_keywords": [],
                    "iocs": [],  # Indicators of Compromise
                    "risk_level": "low"
                },
                "code_statistics": {
                    "total_modules": 0,
                    "total_lines": 0,
                    "total_characters": 0
                }
            }
            
            if vba_analysis["has_macros"]:
                # Extract macro code
                for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
                    if vba_code:
                        module_info = {
                            "filename": filename,
                            "stream_path": stream_path,
                            "vba_filename": vba_filename,
                            "code_length": len(vba_code),
                            "line_count": len(vba_code.splitlines()),
                            "code_preview": vba_code[:500] + "..." if len(vba_code) > 500 else vba_code,
                            "functions": self._extract_vba_functions(vba_code),
                            "subroutines": self._extract_vba_subroutines(vba_code)
                        }
                        vba_analysis["modules"].append(module_info)
                        vba_analysis["code_statistics"]["total_lines"] += module_info["line_count"]
                        vba_analysis["code_statistics"]["total_characters"] += module_info["code_length"]
                
                vba_analysis["code_statistics"]["total_modules"] = len(vba_analysis["modules"])
                
                # Security analysis
                results = vba_parser.analyze_macros()
                if results:
                    for kw_type, keyword, description in results:
                        if kw_type == 'Suspicious':
                            vba_analysis["security_analysis"]["suspicious_keywords"].append({
                                "keyword": keyword,
                                "description": description
                            })
                        elif kw_type == 'AutoExec':
                            vba_analysis["security_analysis"]["auto_exec_keywords"].append({
                                "keyword": keyword,
                                "description": description
                            })
                        elif kw_type == 'IOC':
                            vba_analysis["security_analysis"]["iocs"].append({
                                "keyword": keyword,
                                "description": description
                            })
                
                # Determine risk level
                suspicious_count = len(vba_analysis["security_analysis"]["suspicious_keywords"])
                ioc_count = len(vba_analysis["security_analysis"]["iocs"])
                
                if ioc_count > 0 or suspicious_count > 5:
                    vba_analysis["security_analysis"]["risk_level"] = "high"
                elif suspicious_count > 2:
                    vba_analysis["security_analysis"]["risk_level"] = "medium"
                else:
                    vba_analysis["security_analysis"]["risk_level"] = "low"
            
            vba_parser.close()
            return vba_analysis
            
        except Exception as e:
            return {"error": f"Failed to analyze VBA: {str(e)}"}
    
    def _extract_vba_functions(self, vba_code: str) -> List[str]:
        """Extract function names from VBA code"""
        functions = []
        lines = vba_code.splitlines()
        for line in lines:
            line = line.strip()
            if line.lower().startswith('function ') or line.lower().startswith('public function ') or line.lower().startswith('private function '):
                # Extract function name
                parts = line.split('(')[0].split()
                if len(parts) >= 2:
                    functions.append(parts[-1])
        return functions
    
    def _extract_vba_subroutines(self, vba_code: str) -> List[str]:
        """Extract subroutine names from VBA code"""
        subroutines = []
        lines = vba_code.splitlines()
        for line in lines:
            line = line.strip()
            if line.lower().startswith('sub ') or line.lower().startswith('public sub ') or line.lower().startswith('private sub '):
                # Extract subroutine name
                parts = line.split('(')[0].split()
                if len(parts) >= 2:
                    subroutines.append(parts[-1])
        return subroutines

class MarkdownReportGenerator:
    """Generate comprehensive markdown reports from analysis results"""
    
    def __init__(self, analysis_results: Dict[str, Any]):
        self.results = analysis_results
    
    def generate_report(self) -> str:
        """Generate comprehensive markdown report"""
        report_sections = [
            self._generate_header(),
            self._generate_file_info(),
            self._generate_metadata_section(),
            self._generate_structure_section(),
            self._generate_content_section(),
            self._generate_formatting_section(),
            self._generate_vba_section(),
            self._generate_summary()
        ]
        
        return "\n\n".join(filter(None, report_sections))
    
    def _generate_header(self) -> str:
        """Generate report header"""
        file_name = self.results.get("file_info", {}).get("file_name", "Unknown")
        timestamp = self.results.get("file_info", {}).get("analysis_timestamp", "Unknown")
        
        return f"""# Excel File Analysis Report

**File:** {file_name}  
**Analysis Date:** {timestamp}  
**Generated by:** Excel Analyzer CLI Tool

---"""
    
    def _generate_file_info(self) -> str:
        """Generate file information section"""
        file_info = self.results.get("file_info", {})
        if not file_info:
            return None
        
        return f"""## File Information

| Property | Value |
|----------|-------|
| File Name | {file_info.get('file_name', 'N/A')} |
| File Size | {file_info.get('file_size_mb', 'N/A')} MB ({file_info.get('file_size', 'N/A'):,} bytes) |
| Macro Enabled | {'Yes' if file_info.get('is_macro_enabled') else 'No'} |
| Last Modified | {file_info.get('last_modified', 'N/A')} |"""
    
    def _generate_metadata_section(self) -> str:
        """Generate metadata section"""
        metadata = self.results.get("metadata", {})
        if not metadata or "error" in metadata:
            return "## Metadata\n\n*Metadata analysis failed or unavailable*"
        
        return f"""## Metadata

| Property | Value |
|----------|-------|
| Creator | {metadata.get('creator', 'N/A')} |
| Title | {metadata.get('title', 'N/A')} |
| Subject | {metadata.get('subject', 'N/A')} |
| Created | {metadata.get('created', 'N/A')} |
| Modified | {metadata.get('modified', 'N/A')} |
| Last Modified By | {metadata.get('last_modified_by', 'N/A')} |
| Defined Names | {len(metadata.get('defined_names', []))} |"""
    
    def _generate_structure_section(self) -> str:
        """Generate structure analysis section"""
        structure = self.results.get("structure", {})
        if not structure or "error" in structure:
            return "## Structure Analysis\n\n*Structure analysis failed or unavailable*"
        
        sections = [f"""## Structure Analysis

**Total Sheets:** {structure.get('sheet_count', 0)}

### Sheet Overview"""]
        
        sheets = structure.get("sheets", {})
        for sheet_name, sheet_info in sheets.items():
            protection_info = ""
            if sheet_info.get("protection"):
                prot = sheet_info["protection"]
                if prot.get("sheet_protected"):
                    protection_info = " 🔒"
                if prot.get("password_protected"):
                    protection_info += " 🔐"
            
            sections.append(f"""
#### {sheet_name}{protection_info}

- **Dimensions:** {sheet_info.get('dimensions', 'N/A')}
- **Max Row:** {sheet_info.get('max_row', 0):,}
- **Max Column:** {sheet_info.get('max_column', 0)}
- **Merged Cells:** {sheet_info.get('merged_cells_count', 0)}
- **Has Charts:** {'Yes' if sheet_info.get('has_charts') else 'No'}
- **Has Images:** {'Yes' if sheet_info.get('has_images') else 'No'}""")
        
        return "\n".join(sections)
    
    def _generate_content_section(self) -> str:
        """Generate content analysis section"""
        content = self.results.get("content", {})
        if not content or "error" in content:
            return "## Content Analysis\n\n*Content analysis failed or unavailable*"
        
        sections = [f"""## Content Analysis

**Total Data Rows:** {content.get('total_rows', 0):,}  
**Total Data Columns:** {content.get('total_columns', 0):,}

### Sheet Data Summary"""]
        
        sheets = content.get("sheets", {})
        for sheet_name, sheet_info in sheets.items():
            has_formulas = "Yes" if sheet_info.get("has_formulas") else "No"
            formula_count = sheet_info.get("formula_count_sample", 0)
            
            sections.append(f"""
#### {sheet_name}

- **Rows:** {sheet_info.get('rows', 0):,}
- **Columns:** {sheet_info.get('columns', 0)}
- **Has Formulas:** {has_formulas}
- **Formula Count (sample):** {formula_count}
- **Column Names:** {', '.join(sheet_info.get('column_names', [])[:10])}{'...' if len(sheet_info.get('column_names', [])) > 10 else ''}""")
        
        return "\n".join(sections)
    
    def _generate_formatting_section(self) -> str:
        """Generate formatting analysis section"""
        formatting = self.results.get("formatting", {})
        if not formatting or "error" in formatting:
            return "## Formatting Analysis\n\n*Formatting analysis failed or unavailable*"
        
        summary = formatting.get("summary", {})
        
        sections = [f"""## Formatting Analysis

### Summary
- **Total Styled Cells:** {summary.get('total_styled_cells', 0):,}
- **Unique Fonts:** {len(summary.get('unique_fonts', []))}
- **Unique Colors:** {len(summary.get('unique_colors', []))}
- **Has Conditional Formatting:** {'Yes' if summary.get('has_conditional_formatting') else 'No'}"""]
        
        if summary.get('unique_fonts'):
            sections.append(f"\n**Fonts Used:** {', '.join(summary['unique_fonts'][:10])}{'...' if len(summary['unique_fonts']) > 10 else ''}")
        
        return "\n".join(sections)
    
    def _generate_vba_section(self) -> str:
        """Generate VBA analysis section"""
        vba = self.results.get("vba_analysis", {})
        if not vba:
            return None
        
        if "message" in vba:
            return f"## VBA Analysis\n\n*{vba['message']}*"
        
        if "error" in vba:
            return "## VBA Analysis\n\n*VBA analysis failed or unavailable*"
        
        if not vba.get("has_macros"):
            return "## VBA Analysis\n\n*No VBA macros detected*"
        
        stats = vba.get("code_statistics", {})
        security = vba.get("security_analysis", {})
        
        sections = [f"""## VBA Analysis

### Code Statistics
- **Total Modules:** {stats.get('total_modules', 0)}
- **Total Lines of Code:** {stats.get('total_lines', 0):,}
- **Total Characters:** {stats.get('total_characters', 0):,}

### Security Analysis
- **Risk Level:** {security.get('risk_level', 'unknown').upper()}
- **Suspicious Keywords:** {len(security.get('suspicious_keywords', []))}
- **Auto-Execute Keywords:** {len(security.get('auto_exec_keywords', []))}
- **Indicators of Compromise:** {len(security.get('iocs', []))}"""]
        
        # Add suspicious keywords if any
        if security.get('suspicious_keywords'):
            sections.append("\n### Suspicious Keywords Found")
            for kw in security['suspicious_keywords'][:10]:  # Limit to first 10
                sections.append(f"- **{kw['keyword']}**: {kw['description']}")
        
        # Add module information
        modules = vba.get("modules", [])
        if modules:
            sections.append(f"\n### VBA Modules ({len(modules)})")
            for module in modules[:5]:  # Limit to first 5 modules
                functions = module.get('functions', [])
                subroutines = module.get('subroutines', [])
                sections.append(f"""
#### {module['vba_filename']}
- **Lines:** {module.get('line_count', 0):,}
- **Functions:** {len(functions)} ({', '.join(functions[:5])}{'...' if len(functions) > 5 else ''})
- **Subroutines:** {len(subroutines)} ({', '.join(subroutines[:5])}{'...' if len(subroutines) > 5 else ''})""")
        
        return "\n".join(sections)
    
    def _generate_summary(self) -> str:
        """Generate analysis summary"""
        file_info = self.results.get("file_info", {})
        structure = self.results.get("structure", {})
        content = self.results.get("content", {})
        vba = self.results.get("vba_analysis", {})
        
        summary_points = []
        
        # File complexity
        sheet_count = structure.get("sheet_count", 0)
        total_rows = content.get("total_rows", 0)
        
        if sheet_count > 20:
            summary_points.append("📊 **Complex workbook** with many sheets")
        elif sheet_count > 10:
            summary_points.append("📊 **Medium complexity** workbook")
        else:
            summary_points.append("📊 **Simple workbook** structure")
        
        if total_rows > 10000:
            summary_points.append("📈 **Large dataset** with significant data volume")
        elif total_rows > 1000:
            summary_points.append("📈 **Medium dataset** size")
        
        # VBA analysis
        if vba and vba.get("has_macros"):
            risk_level = vba.get("security_analysis", {}).get("risk_level", "unknown")
            if risk_level == "high":
                summary_points.append("⚠️ **High-risk VBA macros** detected - review recommended")
            elif risk_level == "medium":
                summary_points.append("⚠️ **Medium-risk VBA macros** detected")
            else:
                summary_points.append("✅ **Low-risk VBA macros** detected")
        
        # Formatting complexity
        formatting = self.results.get("formatting", {})
        if formatting and not "error" in formatting:
            styled_cells = formatting.get("summary", {}).get("total_styled_cells", 0)
            if styled_cells > 1000:
                summary_points.append("🎨 **Heavily formatted** workbook")
            elif styled_cells > 100:
                summary_points.append("🎨 **Moderately formatted** workbook")
        
        return f"""## Analysis Summary

{chr(10).join(summary_points) if summary_points else 'Analysis completed successfully.'}

---

*Report generated by Excel Analyzer CLI Tool*  
*For reuse in other applications, extract the analysis data from the JSON output*"""

def main():
    """Main CLI function"""
    parser = argparse.ArgumentParser(
        description="Excel Analyzer - Comprehensive analysis tool for Excel files",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_analyzer.py analyze file.xlsx
  python excel_analyzer.py analyze file.xlsm --include-vba --output-format markdown
  python excel_analyzer.py analyze file.xlsx --output-dir ./reports/
        """
    )
    
    subparsers = parser.add_subparsers(dest='command', help='Available commands')
    
    # Analyze command
    analyze_parser = subparsers.add_parser('analyze', help='Analyze an Excel file')
    analyze_parser.add_argument('file', help='Path to Excel file (.xlsx or .xlsm)')
    analyze_parser.add_argument('--include-vba', action='store_true', default=True,
                               help='Include VBA macro analysis (default: True)')
    analyze_parser.add_argument('--include-formatting', action='store_true', default=True,
                               help='Include formatting analysis (default: True)')
    analyze_parser.add_argument('--output-format', choices=['json', 'markdown', 'both'], 
                               default='both', help='Output format (default: both)')
    analyze_parser.add_argument('--output-dir', default='.', 
                               help='Output directory (default: current directory)')
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        return
    
    if args.command == 'analyze':
        # Validate file exists
        if not os.path.exists(args.file):
            print(f"❌ Error: File '{args.file}' not found")
            return
        
        # Validate file extension
        file_ext = Path(args.file).suffix.lower()
        if file_ext not in ['.xlsx', '.xlsm']:
            print(f"❌ Error: Unsupported file type '{file_ext}'. Only .xlsx and .xlsm files are supported.")
            return
        
        print(f"🚀 Starting analysis of {Path(args.file).name}")
        print("=" * 60)
        
        # Perform analysis
        analyzer = ExcelAnalyzer(args.file)
        results = analyzer.analyze(
            include_vba=args.include_vba,
            include_formatting=args.include_formatting
        )
        
        # Generate outputs
        output_dir = Path(args.output_dir)
        output_dir.mkdir(exist_ok=True)
        
        file_stem = Path(args.file).stem
        
        if args.output_format in ['json', 'both']:
            json_file = output_dir / f"{file_stem}_analysis.json"
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, default=str)
            print(f"📄 JSON report saved: {json_file}")
        
        if args.output_format in ['markdown', 'both']:
            md_generator = MarkdownReportGenerator(results)
            markdown_content = md_generator.generate_report()
            
            md_file = output_dir / f"{file_stem}_analysis.md"
            with open(md_file, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            print(f"📝 Markdown report saved: {md_file}")
        
        print("\n✅ Analysis completed successfully!")
        
        # Print summary
        file_info = results.get("file_info", {})
        structure = results.get("structure", {})
        vba = results.get("vba_analysis", {})
        
        print(f"\n📊 Summary:")
        print(f"   • File size: {file_info.get('file_size_mb', 0)} MB")
        print(f"   • Sheets: {structure.get('sheet_count', 0)}")
        print(f"   • Macro-enabled: {'Yes' if file_info.get('is_macro_enabled') else 'No'}")
        if vba and vba.get("has_macros"):
            print(f"   • VBA modules: {vba.get('code_statistics', {}).get('total_modules', 0)}")
            print(f"   • Security risk: {vba.get('security_analysis', {}).get('risk_level', 'unknown').upper()}")

if __name__ == "__main__":
    main()
