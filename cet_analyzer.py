#!/usr/bin/env python3
"""
CET v22 Excel File Analyzer
Focused analysis for the CET (Cost Estimation Template) v22 file
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime

def analyze_cet_file(file_path: str):
    """Analyze the CET v22 Excel file"""
    print(f"🔍 Analyzing {file_path}...")
    
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    analysis = {
        "file_info": {
            "filename": Path(file_path).name,
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
            "analysis_timestamp": datetime.now().isoformat()
        },
        "sheets": {},
        "summary": {}
    }
    
    print(f"📊 Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")
    
    # Analyze each sheet
    for sheet_name in wb.sheetnames:
        print(f"  📋 Analyzing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        sheet_analysis = analyze_sheet(sheet, sheet_name)
        analysis["sheets"][sheet_name] = sheet_analysis
        
        # Print summary for this sheet
        if sheet_analysis["has_data"]:
            print(f"    ✅ Data found: {sheet_analysis['row_count']} rows, {sheet_analysis['col_count']} columns")
            if sheet_analysis.get("key_fields"):
                print(f"    🔑 Key fields: {', '.join(sheet_analysis['key_fields'][:5])}")
        else:
            print(f"    ⚠️  No data found")
    
    # Generate summary
    analysis["summary"] = generate_summary(analysis["sheets"])
    
    return analysis

def analyze_sheet(sheet, sheet_name: str):
    """Analyze individual sheet content"""
    sheet_analysis = {
        "sheet_name": sheet_name,
        "has_data": False,
        "row_count": 0,
        "col_count": 0,
        "key_fields": [],
        "data_preview": [],
        "structure": {}
    }
    
    # Get sheet dimensions
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    if max_row > 0 and max_col > 0:
        sheet_analysis["has_data"] = True
        sheet_analysis["row_count"] = max_row
        sheet_analysis["col_count"] = max_col
        
        # Get headers (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"Column_{col}")
        
        sheet_analysis["key_fields"] = headers
        
        # Get data preview (first few rows)
        preview_data = []
        for row in range(2, min(6, max_row + 1)):  # First 5 data rows
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            preview_data.append(row_data)
        
        sheet_analysis["data_preview"] = preview_data
        
        # Analyze structure patterns
        sheet_analysis["structure"] = analyze_sheet_structure(sheet, headers, max_row, max_col)
    
    return sheet_analysis

def analyze_sheet_structure(sheet, headers, max_row, max_col):
    """Analyze the structure and patterns in the sheet"""
    structure = {
        "header_patterns": {},
        "data_patterns": {},
        "formulas": [],
        "validation_rules": []
    }
    
    # Analyze header patterns
    for i, header in enumerate(headers):
        if header:
            # Check for common patterns
            if "id" in header.lower():
                structure["header_patterns"]["id_fields"] = structure["header_patterns"].get("id_fields", []) + [i]
            if "name" in header.lower():
                structure["header_patterns"]["name_fields"] = structure["header_patterns"].get("name_fields", []) + [i]
            if "date" in header.lower():
                structure["header_patterns"]["date_fields"] = structure["header_patterns"].get("date_fields", []) + [i]
            if "cost" in header.lower() or "price" in header.lower() or "amount" in header.lower():
                structure["header_patterns"]["financial_fields"] = structure["header_patterns"].get("financial_fields", []) + [i]
            if "status" in header.lower():
                structure["header_patterns"]["status_fields"] = structure["header_patterns"].get("status_fields", []) + [i]
    
    # Check for formulas
    for row in range(2, min(10, max_row + 1)):  # Check first 10 data rows
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.data_type == 'f':  # Formula
                structure["formulas"].append({
                    "cell": f"{cell.column_letter}{row}",
                    "formula": cell.value
                })
    
    return structure

def generate_summary(sheets_analysis):
    """Generate overall summary of the CET file"""
    summary = {
        "total_sheets": len(sheets_analysis),
        "sheets_with_data": 0,
        "total_rows": 0,
        "total_columns": 0,
        "key_sheets": [],
        "estimated_purpose": "",
        "recommended_actions": []
    }
    
    for sheet_name, analysis in sheets_analysis.items():
        if analysis["has_data"]:
            summary["sheets_with_data"] += 1
            summary["total_rows"] += analysis["row_count"]
            summary["total_columns"] = max(summary["total_columns"], analysis["col_count"])
            
            # Identify key sheets
            if analysis["row_count"] > 10:  # Sheets with substantial data
                summary["key_sheets"].append({
                    "name": sheet_name,
                    "rows": analysis["row_count"],
                    "columns": analysis["col_count"],
                    "key_fields": analysis["key_fields"][:5]  # First 5 key fields
                })
    
    # Estimate purpose based on content
    if "cost" in str(sheets_analysis).lower() or "estimate" in str(sheets_analysis).lower():
        summary["estimated_purpose"] = "Cost Estimation Template"
    elif "requirement" in str(sheets_analysis).lower():
        summary["estimated_purpose"] = "Requirements Management"
    else:
        summary["estimated_purpose"] = "Data Template"
    
    # Generate recommendations
    if summary["sheets_with_data"] > 0:
        summary["recommended_actions"].append("File contains structured data suitable for processing")
        summary["recommended_actions"].append("Consider creating data models based on identified patterns")
        summary["recommended_actions"].append("Validate data quality and consistency across sheets")
    
    return summary

def save_analysis(analysis, output_file: str):
    """Save analysis results to file"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, default=str)
    print(f"💾 Analysis saved to: {output_file}")

def main():
    """Main analysis function"""
    file_path = "CET v22.xlsx"
    
    if not Path(file_path).exists():
        print(f"❌ Error: File '{file_path}' not found")
        return
    
    try:
        # Perform analysis
        analysis = analyze_cet_file(file_path)
        
        # Save results
        output_file = "CET_v22_analysis.json"
        save_analysis(analysis, output_file)
        
        # Print summary
        print("\n" + "="*60)
        print("📊 ANALYSIS SUMMARY")
        print("="*60)
        
        summary = analysis["summary"]
        print(f"📁 Total sheets: {summary['total_sheets']}")
        print(f"📊 Sheets with data: {summary['sheets_with_data']}")
        print(f"📈 Total rows: {summary['total_rows']}")
        print(f"📋 Total columns: {summary['total_columns']}")
        print(f"🎯 Estimated purpose: {summary['estimated_purpose']}")
        
        print(f"\n🔑 Key sheets:")
        for sheet in summary["key_sheets"][:3]:  # Show top 3
            print(f"  • {sheet['name']}: {sheet['rows']} rows, {sheet['columns']} columns")
        
        print(f"\n💡 Recommendations:")
        for rec in summary["recommended_actions"]:
            print(f"  • {rec}")
            
    except Exception as e:
        print(f"❌ Error during analysis: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
